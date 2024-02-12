import tkinter as tk
import pandas as pd
import json
import sqlite3
import smtplib
import datetime
import threading
import openpyxl
import os
import ssl

from tkinter import filedialog
from tkinter import ttk
from tkinter import messagebox
from tkcalendar import DateEntry
from email.message import EmailMessage
from credentials import email_password, email_sender

class BookingManagementSystem:
    def __init__(self, root):
        self.root = root
        self.root.title("Booking Management System")

        # Set up a protocol handler to call save_data_to_db before closing
        root.protocol("WM_DELETE_WINDOW", self.on_close)

        # Set the background color to white
        self.root.configure(background='white')

        # Make the window full-screen
        self.root.attributes('-fullscreen', True)

        # Menu Bar
        menu_bar = tk.Menu(root)
        root.config(menu=menu_bar)

        # File Menu
        file_menu = tk.Menu(menu_bar, tearoff=0)
        menu_bar.add_cascade(label="File", menu=file_menu)
        file_menu.add_command(label="Import Data", command=self.import_data)
        file_menu.add_command(label="Export to Excel", command=self.export_to_excel)  # Add Export to Excel option
        file_menu.add_command(label="Update Excel", command=self.update_excel)
        file_menu.add_separator()
        file_menu.add_command(label="Close", command=self.on_close)

        # Search Menu
        search_menu = tk.Menu(menu_bar, tearoff=0)
        menu_bar.add_cascade(label="Search", menu=search_menu)
        search_menu.add_command(label="Filter Data", command=self.show_search_dialog)
        self.revert_filter_enabled = False  # Flag to track whether a filter is applied
        search_menu.add_command(label="Revert Filter", command=self.revert_filter, state=tk.DISABLED)

        self.search_menu = search_menu
        
        # Mail Menu
        mail_menu = tk.Menu(menu_bar, tearoff=0)
        menu_bar.add_cascade(label="Mail", menu=mail_menu)
        mail_menu.add_command(label="Mail to Customer", command=self.open_mail_window)
        mail_menu.add_command(label="Scheduled Mails", command=self.open_scheduled_mails_window)

        # Create a DataFrame for holding booking data
        self.booking_data = pd.DataFrame(columns=['Count', 'Booking Date', 'Travel Date', 'Product', 'Booking Ref', 'Name', 'Country', 'Email', 'Phone No', 'Adult', 'GYG Price', 'Net Price'])

        # Create a table (Treeview) for displaying data
        columns = list(self.booking_data.columns)
        self.tree = ttk.Treeview(root, columns=columns, show='headings')
        self.tree.tag_configure("style", background="white", foreground="black")

        for col in columns:
            self.tree.heading(col, text=col)
            self.tree.column(col, width=100, anchor='center')  # Adjust the width as needed

        for col in self.tree['columns']:
            self.tree.heading(col, text=col, command=lambda c=col: self.on_column_click(c))
            self.tree.column(col, anchor='center')
            col_no_spaces = col.replace(' ', '_')  # Replace spaces with underscores
            if col != 'Count':  # Exclude 'Count' column from B1-Motion and ButtonRelease-1 bindings
                col_index = self.tree['columns'].index(col)
                self.tree.bind('<B1-Motion>', lambda event, c=col, i=col_index: self.on_column_resizing(event, c))
                self.tree.bind(f'<ButtonRelease-1>', self.on_column_release)
                
        # Bind double click event to the treeview
        self.tree.bind('<Double-1>', self.on_row_double_click)

        # Add a Sizegrip widget for automatic column width adjustment
        self.sizegrip = ttk.Sizegrip(root)

        # Configure row and column weights for expanding
        root.grid_rowconfigure(1, weight=1)
        root.grid_columnconfigure(0, weight=1)
        
        self.load_column_configuration()

        # Add a horizontal scrollbar
        x_scrollbar = ttk.Scrollbar(root, orient='horizontal', command=self.tree.xview)
        self.tree.configure(xscrollcommand=x_scrollbar.set)

        # Add a vertical scrollbar
        y_scrollbar = ttk.Scrollbar(root, orient='vertical', command=self.tree.yview)
        self.tree.configure(yscrollcommand=y_scrollbar.set)

        # Place widgets in the grid
        self.tree.pack(fill=tk.BOTH, expand=True)
        x_scrollbar.pack(side=tk.BOTTOM, fill=tk.X)
        y_scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        self.sizegrip.pack(side="bottom", fill="both")

        # SQLite Database connection
        self.db_connection = sqlite3.connect('booking_data.db')
        self.create_table_if_not_exists()

        # Load data from the SQLite database
        self.load_data_from_db()

        self.load_column_configuration()
        
    # Command to convert file to excel
    def export_to_excel(self):
        # Check if there is data to export
        if self.booking_data.empty:
            messagebox.showinfo("No Data", "There is no data to export.")
            return

        # Ask user for file save location
        file_path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")])
        if file_path:
            try:
                # Write data to Excel file
                self.booking_data.to_excel(file_path, index=False)
                messagebox.showinfo("Export Successful", f"Data exported to {file_path} successfully.")
            except Exception as e:
                messagebox.showerror("Export Error", f"An error occurred while exporting data: {e}")
                
    def update_excel(self):
        # Check if there is data to export
        if self.booking_data.empty:
            messagebox.showinfo("No Data", "There is no data to export.")
            return

        # Ask user for file save location
        file_path = filedialog.askopenfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")])
        if file_path:
            try:
                # Read the existing Excel file if it exists
                if os.path.exists(file_path):
                    wb = openpyxl.load_workbook(file_path)
                    ws = wb.active

                    # Get the existing booking references from the Excel sheet
                    existing_booking_refs = set(ws.cell(row=row_index, column=5).value for row_index in range(2, ws.max_row + 1))

                    # Get only the relevant columns from the DataFrame
                    relevant_columns = self.booking_data[['Count', 'Booking Date', 'Travel Date', 'Product', 'Booking Ref', 'Name', 'Country', 'Email', 'Phone No', 'Adult', 'GYG Price', 'Net Price']]

                    # Append new entries to the Excel file
                    for index, row in relevant_columns.iterrows():
                        booking_ref = row['Booking Ref']
                        if booking_ref not in existing_booking_refs:
                            values = row.tolist()
                            ws.append(values)
                            existing_booking_refs.add(booking_ref)

                    # Save the updated Excel file
                    wb.save(file_path)
                    messagebox.showinfo("Update Successful", f"Data updated in {file_path} successfully.")
                else:
                    # If the file doesn't exist, simply export the data
                    self.booking_data[['Count', 'Booking Date', 'Travel Date', 'Product', 'Booking Ref', 'Name', 'Country', 'Email', 'Phone No', 'Adult', 'GYG Price', 'Net Price']].to_excel(file_path, index=False)
                    messagebox.showinfo("Export Successful", f"Data exported to {file_path} successfully.")
            except Exception as e:
                messagebox.showerror("Update Error", f"An error occurred while updating data: {e}")
                
                
    #===============================HANDLE MAIL REQUESTS AND MAIL WINDOWS=======================================#
    def open_scheduled_mails_window(self):
        scheduled_mails_window = tk.Toplevel(self.root)
        scheduled_mails_window.title("Scheduled Mails")

        # Create a frame to display scheduled mails
        scheduled_mails_frame = ttk.Frame(scheduled_mails_window)
        scheduled_mails_frame.pack(fill=tk.BOTH, expand=True)

        # Create a treeview widget to display scheduled mails with the specified columns
        columns = ["Booking Reference", "Customer Name", "Customer Email", "Scheduled Date", "Message"]
        scheduled_mails_tree = ttk.Treeview(scheduled_mails_frame, columns=columns, show='headings')
        scheduled_mails_tree.tag_configure("style", background="white", foreground="black")

        for col in columns:
            scheduled_mails_tree.heading(col, text=col)
            scheduled_mails_tree.column(col, width=100, anchor='center')

        scheduled_mails_tree.pack(fill=tk.BOTH, expand=True)
        
    def open_mail_window(self, booking_ref=None, customer_name=None, customer_email=None, customer_phone=None, customer_travel_date=None):
        def upload_museum_tickets():
            # Function to handle uploading Museum of the Future tickets
            file_path = filedialog.askopenfilename(filetypes=[("All files", "*.*")])
            if file_path:
                museum_tickets_label.config(text=file_path)
                
        def upload_dubai_frame_tickets():
            # Function to handle uploading Dubai Frame tickets
            file_path = filedialog.askopenfilename(filetypes=[("All files", "*.*")])
            if file_path:
                dubai_tickets_label.config(text=file_path)
                
        def send_mail(email_reciever, subject, body, send_datetime=None, **kwargs):
            em = EmailMessage()
            em['From'] = email_sender
            em['To'] = email_reciever
            em['Subject'] = subject
            em.set_content(body)
            
            context = ssl.create_default_context()
            
            attachments = kwargs.get('attachments', None)
                
            # Attach files if any
            if attachments:
                for attachment in attachments:
                    with open(attachment, 'rb') as file:
                        file_data = file.read()
                        file_name = os.path.basename(attachment)
                    em.add_attachment(file_data, maintype='application', subtype='octet-stream', filename=file_name)
                    
            with smtplib.SMTP_SSL('smtp.gmail.com', 465, context=context) as smtp:
                smtp.login(email_sender, email_password)
                smtp.sendmail(email_sender, email_reciever, em.as_string())
                
        def fetch_data():
            booking_ref = booking_ref_entry.get()
            if booking_ref:
                # Connect to the database
                conn = sqlite3.connect('booking_data.db')
                conn.row_factory = sqlite3.Row  # Set row_factory to get rows as dictionaries

                # Create a cursor
                cursor = conn.cursor()

                # Execute the query
                cursor.execute("SELECT Name, Email, [Phone No], [Travel Date] FROM bookings WHERE [Booking Ref]=?", (booking_ref,))

                # Fetch one row
                row = cursor.fetchone()

                if row:
                    # Autofill customer name, email, phone, and travel date fields
                    customer_name_entry.delete(0, tk.END)
                    customer_name_entry.insert(0, row['Name'])
                    
                    customer_mail_entry.delete(0, tk.END)
                    customer_mail_entry.insert(0, row['Email'])

                    customer_phone_entry.delete(0, tk.END)
                    customer_phone_entry.insert(0, row['Phone No'])

                    customer_travel_date_entry.delete(0, tk.END)
                    customer_travel_date_entry.insert(0, row['Travel Date'])
                else:
                    messagebox.showerror("Error", "No data found for the given booking reference.")

                # Close the cursor and connection
                cursor.close()
                conn.close()
            else:
                messagebox.showerror("Error", "Please enter a booking reference.")

        mail_window = tk.Toplevel(self.root)
        mail_window.title("Send Mail to Customer")

        # Labels and entry widgets for booking reference, mail subject, mail body
        tk.Label(mail_window, text="Booking Reference:").grid(row=0, column=0, padx=10, pady=10)
        booking_ref_entry = tk.Entry(mail_window)
        booking_ref_entry.grid(row=0, column=1, padx=10, pady=10)

        # Button to fetch data
        fetch_button = tk.Button(mail_window, text="Fetch Data", command=fetch_data)
        fetch_button.grid(row=1, column=0, columnspan=2, padx=10, pady=10)

        # Labels for customer mail, phone, and travel date
        tk.Label(mail_window, text="Customer Name:").grid(row=2, column=0, padx=10, pady=10)
        customer_name_entry = tk.Entry(mail_window)
        customer_name_entry.grid(row=2, column=1, padx=10, pady=10)
        
        tk.Label(mail_window, text="Customer Mail:").grid(row=3, column=0, padx=10, pady=10)
        customer_mail_entry = tk.Entry(mail_window)
        customer_mail_entry.grid(row=3, column=1, padx=10, pady=10)

        tk.Label(mail_window, text="Customer Phone:").grid(row=4, column=0, padx=10, pady=10)
        customer_phone_entry = tk.Entry(mail_window)
        customer_phone_entry.grid(row=4, column=1, padx=10, pady=10)

        tk.Label(mail_window, text="Customer Travel Date:").grid(row=5, column=0, padx=10, pady=10)
        customer_travel_date_entry = tk.Entry(mail_window)
        customer_travel_date_entry.grid(row=5, column=1, padx=10, pady=10)

        tk.Label(mail_window, text="Mail Subject:").grid(row=6, column=0, padx=10, pady=10)
        mail_subject_entry = tk.Entry(mail_window)
        mail_subject_entry.grid(row=6, column=1, padx=10, pady=10)

        tk.Label(mail_window, text="Mail Body:").grid(row=7, column=0, padx=10, pady=10)
        mail_body_entry = tk.Text(mail_window, height=5, width=30)
        mail_body_entry.grid(row=7, column=1, padx=10, pady=10)

        # File upload options
        tk.Label(mail_window, text="Museum of the Future Tickets:").grid(row=8, column=0, padx=10, pady=10)
        museum_tickets_button = tk.Button(mail_window, text="Upload", command=upload_museum_tickets)
        museum_tickets_button.grid(row=8, column=1, padx=10, pady=10)
        museum_tickets_label = tk.Label(mail_window, text="")
        museum_tickets_label.grid(row=8, column=2, padx=10, pady=10)

        tk.Label(mail_window, text="Dubai Frame Tickets:").grid(row=9, column=0, padx=10, pady=10)
        dubai_tickets_button = tk.Button(mail_window, text="Upload", command=upload_dubai_frame_tickets)
        dubai_tickets_button.grid(row=9, column=1, padx=10, pady=10)
        dubai_tickets_label = tk.Label(mail_window, text="")
        dubai_tickets_label.grid(row=9, column=2, padx=10, pady=10)
        
        # Label and DateEntry widget for selecting date
        tk.Label(mail_window, text="Schedule Mail Date:").grid(row=10, column=0, padx=10, pady=10)
        schedule_date_entry = DateEntry(mail_window, date_pattern='dd-mm-yyyy')
        schedule_date_entry.grid(row=10, column=1, padx=10, pady=10)

        # Label and Entry widget for selecting time
        tk.Label(mail_window, text="Schedule Mail Time (24-hour format):").grid(row=11, column=0, padx=10, pady=10)
        schedule_time_entry = tk.Entry(mail_window)
        schedule_time_entry.grid(row=11, column=1, padx=10, pady=10)
        
        def schedule_and_send_mail():
            # Get selected date and time
            send_date = schedule_date_entry.get_date()
            send_time_str = schedule_time_entry.get()
            send_time = datetime.datetime.strptime(send_time_str, "%H:%M").time()

            # Combine date and time
            send_datetime = datetime.datetime.combine(send_date, send_time)

            # Schedule the mail sending task
            threading.Thread(target=send_mail_threaded, args=(send_datetime,)).start()
            
        def send_mail_threaded(send_datetime):
            # Get other mail details
            email_receiver = customer_mail_entry.get()
            subject = mail_subject_entry.get()
            body = mail_body_entry.get('1.0', 'end')
            attachments = [museum_tickets_label.cget("text"), dubai_tickets_label.cget("text")]

            # Calculate the delay in seconds
            delay = (send_datetime - datetime.datetime.now()).total_seconds()

            # If the delay is negative, send the mail immediately
            if delay <= 0:
                send_mail(email_receiver, subject, body, attachments=attachments)
            else:
                # Wait for the specified time before sending the mail
                threading.Timer(delay, send_mail, args=(email_receiver, subject, body), kwargs={"attachments": attachments}).start()

        
        # Button to send mail
        send_button = tk.Button(mail_window, text="Schedule and Send Mail", command=schedule_and_send_mail)
        send_button.grid(row=12, column=0, columnspan=2, padx=10, pady=10)
        
        # Set default values for entry fields based on the extracted details
        booking_ref_entry.insert(0, booking_ref)
        customer_name_entry.insert(0, customer_name)
        customer_mail_entry.insert(0, customer_email)
        customer_phone_entry.insert(0, customer_phone)
        customer_travel_date_entry.insert(0, customer_travel_date)
        
    def on_row_double_click(self, event):
        # Get the selected item from the event
        item = self.tree.selection()[0]
        
        # Get the data of the selected row
        selected_row = self.tree.item(item, 'values')
        
        # print(selected_row)  # Add this line to print the selected row
        
        # Extract relevant information from the selected row
        booking_ref = selected_row[3]  # Booking Reference
        customer_name = selected_row[4]  # Customer Name
        customer_email = selected_row[5]  # Customer Email
        customer_phone = selected_row[6]  # Customer Phone No
        customer_travel_date = selected_row[2]  # Customer Travel Date

        # Open the mail window with the extracted details
        self.open_mail_window(booking_ref, customer_name, customer_email, customer_phone, customer_travel_date)

    #======================================DB CONNECTIONS AND CONFIGURATIONS====================================#
        
    def on_close(self):
        # Save data to the SQLite database before closing
        self.save_data_to_db()
        self.load_column_configuration()
        self.root.destroy()
    
    def create_table_if_not_exists(self):
        # Create a table if it doesn't exist in the SQLite database
        query = '''
        CREATE TABLE IF NOT EXISTS bookings (
            ID INTEGER PRIMARY KEY AUTOINCREMENT,
            Booking_Date TEXT,
            Travel_Date TEXT,
            Product TEXT,
            Booking_Ref TEXT,
            Name TEXT,
            Country TEXT,
            Phone_No TEXT,
            Adult INTEGER,
            GYG_Price REAL,
            Net_Price REAL,
            Email TEXT
        );
        '''
        self.db_connection.execute(query)
        self.db_connection.commit()

    def save_data_to_db(self):
        # Save data to the SQLite database
        self.db_connection.execute('DELETE FROM bookings;')  # Clear existing data
        self.booking_data.to_sql('bookings', self.db_connection, index=False, if_exists='replace')
        self.db_connection.commit()

    def load_data_from_db(self):
        # Load data from the SQLite database
        try:
            self.booking_data = pd.read_sql_query('SELECT * FROM bookings;', self.db_connection)
            self.update_treeview()
        except pd.io.sql.DatabaseError:
            # Use default data if the table is empty or not found
            self.booking_data = pd.DataFrame(columns=['Booking_Date', 'Travel_Date', 'Product', 'Booking_Ref', 'Name', 'Country', 'Phone_No', 'Adult', 'GYG_Price', 'Net_Price', 'Email'])
            self.update_treeview()

    def import_data(self):
        file_path = filedialog.askopenfilename(filetypes=[('Excel Files', '*.xlsx;*.xls')])
        if file_path:
            try:
                all_data = pd.read_excel(file_path, sheet_name=None)

                # Clear existing data in the DataFrame and the SQLite database
                self.booking_data = pd.DataFrame(columns=['Booking_Date', 'Travel_Date', 'Product', 'Booking_Ref', 'Name', 'Country', 'Phone_No', 'Adult', 'GYG_Price', 'Net_Price', 'Email'])
                self.save_data_to_db()

                count = 0

                # Iterate through all sheets and append data to the DataFrame and SQLite database
                for sheet_name, data in all_data.items():
                    sheet_data = pd.DataFrame()
                    sheet_data['Booking_Date'] = pd.to_datetime(data['Purchase Date (local time)'], format='%d/%m/%Y %H:%M', errors='coerce').dt.strftime('%d/%m/%Y %H:%M')
                    sheet_data['Travel_Date'] = pd.to_datetime(data['Date'], format='%d/%m/%Y', errors='coerce').dt.strftime('%d/%m/%Y')
                    sheet_data['Product'] = data['Product']
                    sheet_data['Booking_Ref'] = data['Booking Ref #']
                    sheet_data['Name'] = data['Traveler\'s First Name'] + ' ' + data['Traveler\'s Last Name']
                    sheet_data['Country'] = data['Traveler\'s Country']
                    sheet_data['Phone_No'] = data['Phone']
                    sheet_data['Adult'] = pd.to_numeric(data['Adult'], errors='coerce')
                    sheet_data['GYG_Price'] = pd.to_numeric(data['Price'].str.replace(' AED', ''), errors='coerce')
                    sheet_data['Net_Price'] = pd.to_numeric(data['Net Price'].str.replace(' AED', ''), errors='coerce')
                    sheet_data['Email'] = data['Email']
                    sheet_data['Count'] = range(count + 1, count + 1 + len(sheet_data))
                    count += len(sheet_data)

                    # Append data to the main DataFrame and SQLite database
                    self.booking_data = pd.concat([self.booking_data, sheet_data], ignore_index=True)
                    self.save_data_to_db()

                print("Data Imported and Transformed Successfully!")
            except Exception as e:
                print(f"Error importing and transforming data: {e}")

    #=========================================COLUMN RESIZING===================================================#

    def on_column_click(self, column):
        # Save the current column configuration before resizing starts
        self.save_column_configuration()

    def on_column_resizing(self, event, column):
        # Update the configuration while the user is actively resizing
        original_width = self.tree.column(column, 'width')
        delta_x = event.x - event.x_root
        new_width = max(original_width + delta_x, 1)  # Ensure the width is not negative

        # Calculate the width change relative to the original width
        delta_width = new_width - original_width

        # Update the width of the current column
        self.tree.column(column, width=new_width)

        # Spread the width change among other columns
        for col in self.tree['columns']:
            if col != column:
                current_width = self.tree.column(col, 'width')
                self.tree.column(col, width=current_width - delta_width)

    def on_column_release(self, event):
        # Save the column configuration after resizing is done
        self.save_column_configuration()

    def save_column_configuration(self):
        # Save the column configuration to a file (e.g., config.json)
        config = {col: {'width': self.tree.column(col, 'width')} for col in self.tree['columns']}
        with open('config.json', 'w') as file:
            json.dump(config, file)

    def load_column_configuration(self):
        try:
            # Load the column configuration from a file (e.g., config.json)
            with open('config.json', 'r') as file:
                config = json.load(file)

            # Apply the saved column widths
            for col, values in config.items():
                self.tree.column(col, width=values['width'], anchor='center')

        except FileNotFoundError:
            # Use default column configuration if the file is not found
            pass
        except Exception as e:
            print(f"Error loading column configuration: {e}")

    #===================================================IMPORT DATA====================================================#

    def import_data(self):
        file_path = filedialog.askopenfilename(filetypes=[('Excel Files', '*.xlsx;*.xls')])
        if file_path:
            try:
                all_data = pd.read_excel(file_path, sheet_name=None)

                # Clear existing data in the DataFrame
                self.booking_data = pd.DataFrame(columns=['Count', 'Booking Date', 'Travel Date', 'Product', 'Booking Ref', 'Name', 'Country', 'Email'])

                count = 0

                # Iterate through all sheets and append data to the DataFrame
                for sheet_name, data in all_data.items():
                    sheet_data = pd.DataFrame()
                    sheet_data['Booking Date'] = pd.to_datetime(data['Purchase Date (local time)'], format='%d/%m/%Y %H:%M', errors='coerce').dt.strftime('%d/%m/%Y %H:%M')
                    sheet_data['Travel Date'] = pd.to_datetime(data['Date'], format='%d/%m/%Y', errors='coerce').dt.strftime('%d/%m/%Y')
                    sheet_data['Product'] = data['Product']
                    sheet_data['Booking Ref'] = data['Booking Ref #']
                    sheet_data['Name'] = data['Traveler\'s First Name'] + ' ' + data['Traveler\'s Last Name']
                    sheet_data['Country'] = data['Traveler\'s Country']
                    sheet_data['Phone No'] = data['Phone']
                    sheet_data['Adult'] = pd.to_numeric(data['Adult'], errors='coerce')
                    sheet_data['GYG Price'] = pd.to_numeric(data['Price'].str.replace(' AED', ''), errors='coerce')
                    sheet_data['Net Price'] = pd.to_numeric(data['Net Price'].str.replace(' AED', ''), errors='coerce')
                    sheet_data['Email'] = data['Email']
                    sheet_data['Count'] = range(count + 1, count + 1 + len(sheet_data))
                    count += len(sheet_data)

                    # Append data to the main DataFrame
                    self.booking_data = pd.concat([self.booking_data, sheet_data], ignore_index=True)
                    

                self.update_treeview()
                print("Data Imported and Transformed Successfully!")
            except Exception as e:
                print(f"Error importing and transforming data: {e}")

    #================================================SEACRH DATA====================================================#
    def show_search_dialog(self):
        # Create a search dialog
        search_dialog = tk.Toplevel(self.root)
        search_dialog.title("Search Options")

        # Create a label and dropdown menu for selecting search criteria
        tk.Label(search_dialog, text="Select Search Criteria:").grid(row=0, column=0, padx=10, pady=10)
        search_criteria_var = tk.StringVar()
        search_criteria_var.set("Booking Ref")  # Set default value
        search_criteria_menu = ttk.Combobox(search_dialog, textvariable=search_criteria_var, values=['Booking Ref', 'Customer Name', 'No of Adults', 'Travel Date'])
        search_criteria_menu.grid(row=0, column=1, padx=10, pady=10)

        # Create a label and entry widget for entering search value
        tk.Label(search_dialog, text="Enter Search Value:").grid(row=1, column=0, padx=10, pady=10)
        search_value_var = tk.StringVar()
        search_value_entry = tk.Entry(search_dialog, textvariable=search_value_var)
        search_value_entry.grid(row=1, column=1, padx=10, pady=10)

        # Create a date picker for selecting the date
        tk.Label(search_dialog, text="Select Date:").grid(row=2, column=0, padx=10, pady=10)
        date_var = tk.StringVar()
        date_entry = DateEntry(search_dialog, textvariable=date_var, date_pattern='dd/mm/yyyy')
        date_entry.grid(row=2, column=1, padx=10, pady=10)
        date_entry.grid_remove()  # Hide the date picker initially

        # Function to toggle between entry and date picker
        def toggle_input_widget():
            if search_criteria_var.get() == 'Travel Date':
                search_value_entry.grid_remove()
                date_entry.grid()
            else:
                date_entry.grid_remove()
                search_value_entry.grid()

        # Bind the function to the search criteria dropdown
        search_criteria_menu.bind("<<ComboboxSelected>>", lambda event: toggle_input_widget())

        # Create a button to apply the search
        search_button = tk.Button(search_dialog, text="Search", command=lambda: self.apply_search(search_criteria_var.get(), search_value_var.get(), date_var.get(), search_dialog))
        search_button.grid(row=3, column=0, columnspan=2, pady=10)

        # Enable or disable "Revert Filter" based on the filter status
        self.search_menu.entryconfig("Revert Filter", state=tk.NORMAL if self.revert_filter_enabled else tk.DISABLED)

    def apply_search(self, criteria, value, date_value, search_dialog):
        # Apply the search and update the treeview
        if criteria == 'Booking Ref':
            result = self.booking_data[self.booking_data['Booking Ref'].astype(str).str.contains(value, case=False)]
        elif criteria == 'Customer Name':
            result = self.booking_data[self.booking_data['Name'].astype(str).str.contains(value, case=False)]
        elif criteria == 'No of Adults':
            result = self.booking_data[self.booking_data['Adult'] == int(value)]
        elif criteria == 'Travel Date':
            result = self.booking_data[self.booking_data['Travel Date'] == date_value]

        # Update the treeview with the search result
        self.update_treeview(data=result)

        # Set the flag to indicate that a filter is applied
        self.revert_filter_enabled = True

        # Enable "Revert Filter" based on the filter status
        self.search_menu.entryconfig("Revert Filter", state=tk.NORMAL)

        # Destroy the search dialog
        search_dialog.destroy()

    def revert_filter(self):
        # Revert the filter and update the treeview
        self.update_treeview(data=self.booking_data)

        # Reset the flag to indicate that no filter is applied
        self.revert_filter_enabled = False

        # Disable "Revert Filter" since no filter is applied
        self.search_menu.entryconfig("Revert Filter", state=tk.DISABLED)

    def update_treeview(self, data=None):
        # Clear existing items in the treeview
        for item in self.tree.get_children():
            self.tree.delete(item)

        # Insert new data into the treeview
        if data is None:
            data = self.booking_data

        for index, row in data.iterrows():
            values = row.tolist()
            values[-1] = int(values[-1])  # Convert 'Adult' column to integer without decimal points
            values[-2] = str(values[-2]) if pd.notna(values[-2]) else ""  # Convert 'Phone No' column to string
            self.tree.insert('', index, values=row.tolist(), tags=("style",))

        self.save_column_configuration()

if __name__ == "__main__":
    root = tk.Tk()
    app = BookingManagementSystem(root)
    root.mainloop()